"""
analisis_segmentacion.py
─────────────────────────────────────────────────────────────────────────────
Análisis exploratorio + segmentación de mercado + regresión (departamentos)
sobre leads_lamudi_enriched_openai.csv

Segmentos:
  1. Estudiantes        — precio bajo, amueblado, cerca de campus
  2. Tercera edad       — elevador / PB, seguridad, tranquilo
  3. Extranjeros/Chinos — amueblado, amenidades premium, colonia top, precio alto

Output:
  figures/seg_*.png           — gráficas
  segmentos_lamudi.xlsx       — Excel con una hoja por segmento + leads
  segmentos_lamudi_ranks.csv  — ranking consolidado los tres segmentos
─────────────────────────────────────────────────────────────────────────────
"""

import re
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats

from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import Ridge
from sklearn.model_selection import cross_val_score, KFold, train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, mean_squared_error
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings('ignore')
sns.set_theme(style='whitegrid', palette='muted', font_scale=1.1)

BLUE   = '#2C5F8A'
GREEN  = '#27AE60'
ORANGE = '#F39C12'
RED    = '#E74C3C'
OUT    = './figures/'
CSV_IN = './leads_lamudi_enriched_openai.csv'

# ══════════════════════════════════════════════════════════════════════════
# CARGA
# ══════════════════════════════════════════════════════════════════════════
print('Cargando datos...')
df = pd.read_csv(CSV_IN)
print(f'  {len(df):,} registros  |  {len(df.columns)} columnas\n')

desc_lower = df['description'].fillna('').str.lower()

# ══════════════════════════════════════════════════════════════════════════
# 1. EDA EXPLORATORIO
# ══════════════════════════════════════════════════════════════════════════
print('── 1. EDA ──')

# 1a. Distribución de precio por tipo de propiedad
fig, ax = plt.subplots(figsize=(12, 5))
tipos_plot = ['departamento', 'casa', 'oficina', 'local_comercial']
data_box   = [df[df['property_type'] == t]['price'].dropna() for t in tipos_plot]
bp = ax.boxplot(data_box, labels=tipos_plot, patch_artist=True,
                showfliers=False,
                boxprops=dict(facecolor='#AED6F1'),
                medianprops=dict(color=BLUE, linewidth=2))
ax.set_title('Distribución de precio por tipo de propiedad\n(sin outliers extremos)', fontsize=13)
ax.set_ylabel('Precio mensual (MXN)')
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
plt.tight_layout()
plt.savefig(OUT + 'seg_01_precio_por_tipo.png', dpi=150)
plt.close()
print('  ✓ seg_01_precio_por_tipo.png')

# 1b. Top 15 colonias por precio mediano (departamentos)
dep = df[df['property_type'] == 'departamento'].copy()
top_col = (dep.groupby('colonia')['price']
             .agg(['median', 'count'])
             .query('count >= 3')
             .sort_values('median', ascending=True)
             .tail(15))

fig, ax = plt.subplots(figsize=(10, 6))
bars = ax.barh(top_col.index, top_col['median'], color=BLUE)
for bar, (_, row) in zip(bars, top_col.iterrows()):
    ax.text(bar.get_width() + 500, bar.get_y() + bar.get_height()/2,
            f'n={int(row["count"])}', va='center', fontsize=8)
ax.set_title('Top 15 colonias — Precio mediano (Departamentos)', fontsize=13)
ax.set_xlabel('Precio mediano mensual (MXN)')
ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
plt.tight_layout()
plt.savefig(OUT + 'seg_02_colonias_precio.png', dpi=150)
plt.close()
print('  ✓ seg_02_colonias_precio.png')

# 1c. Frecuencia de amenidades
amenidades = {
    'Seguridad 24h':  'has_security_24h',
    'Minisplit/AC':   'has_minisplit_ac',
    'Terraza':        'has_terrace',
    'Elevador':       'has_elevator',
    'Alberca':        'has_pool',
    'Gimnasio':       'has_gym',
    'Sala de juntas': 'has_sala_juntas',
    'Bodega':         'has_bodega',
    'Lobby':          'has_lobby',
    'Coworking':      'has_coworking',
    'Jardín':         'has_garden',
    'Calefacción':    'has_heating',
    'Permite mascotas': 'allows_pets',
    'Cuarto servicio':  'has_service_room',
}
amenidad_pct = {k: (df[v] == 1).sum() / len(df) * 100
                for k, v in amenidades.items()}
amenidad_s = pd.Series(amenidad_pct).sort_values()

fig, ax = plt.subplots(figsize=(9, 6))
colors_a = [GREEN if v > 15 else BLUE for v in amenidad_s.values]
ax.barh(amenidad_s.index, amenidad_s.values, color=colors_a)
ax.set_title('Frecuencia de amenidades — todos los tipos', fontsize=13)
ax.set_xlabel('% de propiedades')
for i, v in enumerate(amenidad_s.values):
    ax.text(v + 0.3, i, f'{v:.1f}%', va='center', fontsize=9)
plt.tight_layout()
plt.savefig(OUT + 'seg_03_amenidades.png', dpi=150)
plt.close()
print('  ✓ seg_03_amenidades.png')

# ══════════════════════════════════════════════════════════════════════════
# 2. SEGMENTACIÓN DE MERCADO
# ══════════════════════════════════════════════════════════════════════════
print('\n── 2. Segmentación ──')

# ── Palabras clave por segmento ───────────────────────────────────────────
KW_ESTUDIANTES = r'estudiante|universitari|tec |tecnológico|tecnologico|itesm|uanl|udem|campus|college'
KW_TERCERA_EDAD = r'tranquilo|tranquila|silencio|seguro|segura|privada|privado|adulto mayor|pensionado'
KW_EXTRANJERO = r'ejecutivo|corporativo|expat|relocation|internacional|vista|penthouse|premium|exclusivo|lujo'

COLONIAS_PREMIUM = {
    'Valle Oriente', 'Obispado', 'San Pedro Garza García', 'Del Paseo Residencial',
    'Colinas de San Jerónimo', 'Cumbres', 'Lomas de San Francisco',
    'Cumbres del Sol', 'Loma Larga', 'Contry',
}

# ── Score Estudiantes ─────────────────────────────────────────────────────
def score_estudiante(row, desc):
    s = 0
    if row['price'] <= 12000:   s += 3
    elif row['price'] <= 18000: s += 2
    elif row['price'] <= 25000: s += 1
    if row['furnished'] == 1:   s += 3
    if row['property_type'] in ('departamento', 'casa'): s += 2
    if re.search(KW_ESTUDIANTES, desc):  s += 4
    if row['has_minisplit_ac'] == 1:     s += 1
    if row['bedrooms'] in (1, 2):        s += 1
    if row['allows_pets'] == 1:          s += 1
    return s

# ── Score Tercera Edad ────────────────────────────────────────────────────
def score_tercera_edad(row, desc):
    s = 0
    if row['has_elevator'] == 1:     s += 4
    if row['floor'] == 0:            s += 3  # planta baja
    if row['has_security_24h'] == 1: s += 3
    if row['has_lobby'] == 1:        s += 1
    if row['property_type'] in ('departamento', 'casa'): s += 2
    if re.search(KW_TERCERA_EDAD, desc): s += 3
    if row['price'] <= 30000:            s += 2
    if row['has_gym'] == 1:              s += 1
    if row['has_garden'] == 1:           s += 1
    if row['furnished'] == 1:            s += 1
    return s

# ── Score Extranjero / Chino ──────────────────────────────────────────────
def score_extranjero(row, desc):
    s = 0
    if row['furnished'] == 1:        s += 4
    if row['has_security_24h'] == 1: s += 3
    if row['has_gym'] == 1:          s += 2
    if row['has_pool'] == 1:         s += 2
    if row['has_elevator'] == 1:     s += 2
    if row['has_lobby'] == 1:        s += 1
    if row['has_sala_juntas'] == 1:  s += 1
    if row['has_coworking'] == 1:    s += 1
    if row['property_type'] in ('departamento', 'oficina'): s += 1
    if row['colonia'] in COLONIAS_PREMIUM: s += 3
    if re.search(KW_EXTRANJERO, desc): s += 3
    if row['price'] >= 25000:          s += 2
    if row['price'] >= 50000:          s += 1
    return s

df['score_estudiante']   = [score_estudiante(r, d)   for r, d in zip(df.to_dict('records'), desc_lower)]
df['score_tercera_edad'] = [score_tercera_edad(r, d) for r, d in zip(df.to_dict('records'), desc_lower)]
df['score_extranjero']   = [score_extranjero(r, d)   for r, d in zip(df.to_dict('records'), desc_lower)]

# ── Top 30 por segmento ───────────────────────────────────────────────────
COLS_LEADS = ['name', 'phone', 'source_url', 'address', 'property_type', 'price',
              'colonia', 'area_m2', 'bedrooms', 'bathrooms', 'parking',
              'furnished', 'has_elevator', 'has_security_24h', 'has_gym',
              'has_pool', 'amenidades_count', 'notes']

top_est  = df.nlargest(30, 'score_estudiante')[COLS_LEADS + ['score_estudiante']].copy()
top_3ra  = df.nlargest(30, 'score_tercera_edad')[COLS_LEADS + ['score_tercera_edad']].copy()
top_ext  = df.nlargest(30, 'score_extranjero')[COLS_LEADS + ['score_extranjero']].copy()

for seg, top, scol in [
    ('Estudiantes', top_est, 'score_estudiante'),
    ('Tercera edad', top_3ra, 'score_tercera_edad'),
    ('Extranjeros', top_ext, 'score_extranjero'),
]:
    print(f'  {seg:<15}: score máx={top[scol].max()}  '
          f'precio mediano=${top["price"].median():,.0f}  '
          f'con teléfono={top["phone"].notna().sum()}')

# ── Gráfica scores ────────────────────────────────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(15, 5))
for ax, (seg, col, color) in zip(axes, [
    ('Estudiantes',  'score_estudiante',   BLUE),
    ('Tercera edad', 'score_tercera_edad', GREEN),
    ('Extranjeros',  'score_extranjero',   ORANGE),
]):
    ax.hist(df[col], bins=20, color=color, edgecolor='white')
    ax.axvline(df[col].quantile(0.9), color=RED, linestyle='--',
               label=f'p90 = {df[col].quantile(0.9):.0f}')
    ax.set_title(f'Distribución score\n{seg}')
    ax.set_xlabel('Score')
    ax.set_ylabel('# propiedades')
    ax.legend()
plt.suptitle('Distribución de scores por segmento', fontsize=13)
plt.tight_layout()
plt.savefig(OUT + 'seg_04_scores.png', dpi=150)
plt.close()
print('\n  ✓ seg_04_scores.png')

# ── Precio vs score scatter ───────────────────────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(15, 5))
for ax, (seg, col, color) in zip(axes, [
    ('Estudiantes',  'score_estudiante',   BLUE),
    ('Tercera edad', 'score_tercera_edad', GREEN),
    ('Extranjeros',  'score_extranjero',   ORANGE),
]):
    sub = df[df['price'] < df['price'].quantile(0.95)]
    ax.scatter(sub[col], sub['price'], alpha=0.3, color=color, s=15)
    ax.set_xlabel('Score')
    ax.set_ylabel('Precio (MXN)')
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${x:,.0f}'))
    ax.set_title(f'{seg}\nr={sub[col].corr(sub["price"]):.2f}')
plt.suptitle('Precio vs Score por segmento', fontsize=13)
plt.tight_layout()
plt.savefig(OUT + 'seg_05_precio_vs_score.png', dpi=150)
plt.close()
print('  ✓ seg_05_precio_vs_score.png')

# ══════════════════════════════════════════════════════════════════════════
# 3. REGRESIÓN — DEPARTAMENTOS
# ══════════════════════════════════════════════════════════════════════════
print('\n── 3. Regresión (departamentos) ──')

dep = df[df['property_type'] == 'departamento'].copy()

# Target encoding de colonia
loc_med = dep.groupby('colonia')['price'].median()
dep['colonia_tgt'] = dep['colonia'].map(loc_med).fillna(dep['price'].median())

FEATURES = [
    'area_m2', 'bedrooms', 'bathrooms', 'parking', 'floor',
    'furnished', 'is_new_construction',
    'has_pool', 'has_gym', 'has_elevator', 'has_terrace',
    'has_security_24h', 'has_minisplit_ac', 'has_service_room',
    'amenidades_count', 'colonia_tgt',
]

dep_m = dep.dropna(subset=['area_m2', 'price']).copy()
dep_m[FEATURES] = dep_m[FEATURES].fillna(0)
mask  = dep_m[FEATURES + ['price']].notna().all(axis=1)
dep_m = dep_m[mask]

print(f'  Departamentos con area_m2: {len(dep_m)} de {len(dep)}')

if len(dep_m) < 30:
    print('  ⚠ Muy pocos registros completos para regresión confiable.')
else:
    X = dep_m[FEATURES].values
    y = dep_m['price'].values

    # Winsorizar precio al p95
    p95 = np.percentile(y, 95)
    mask_w = y <= p95
    X, y = X[mask_w], y[mask_w]
    print(f'  Tras winsorizar al p95: {len(y)} registros')

    X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2, random_state=42)
    kf = KFold(n_splits=5, shuffle=True, random_state=42)

    models = {
        'Ridge':         Ridge(alpha=1.0),
        'Random Forest': RandomForestRegressor(n_estimators=300, random_state=42, n_jobs=-1),
        'Gradient Boost': GradientBoostingRegressor(n_estimators=200, random_state=42),
    }

    results = {}
    for name, model in models.items():
        model.fit(X_tr, y_tr)
        y_pred = model.predict(X_te)
        mae  = mean_absolute_error(y_te, y_pred)
        r2   = r2_score(y_te, y_pred)
        cv   = cross_val_score(model, X, y, cv=kf, scoring='r2').mean()
        results[name] = {'MAE': mae, 'R²': r2, 'CV-R²': cv}
        print(f'  {name:<18}: R²={r2:.3f}  CV-R²={cv:.3f}  MAE=${mae:,.0f}')

    res_df = pd.DataFrame(results).T
    best   = res_df['R²'].idxmax()

    # Feature importance
    rf = models['Random Forest']
    imp = pd.Series(rf.feature_importances_, index=FEATURES).sort_values().tail(12)
    fig, ax = plt.subplots(figsize=(9, 5))
    colors_i = [GREEN if f in ('area_m2','bedrooms','bathrooms') else BLUE for f in imp.index]
    ax.barh(imp.index, imp.values, color=colors_i)
    ax.set_title(f'Feature importance — Random Forest (Departamentos)\nR²={results["Random Forest"]["R²"]:.3f}  CV-R²={results["Random Forest"]["CV-R²"]:.3f}', fontsize=12)
    ax.set_xlabel('Importancia relativa')
    plt.tight_layout()
    plt.savefig(OUT + 'seg_06_feature_importance.png', dpi=150)
    plt.close()
    print('  ✓ seg_06_feature_importance.png')

# ══════════════════════════════════════════════════════════════════════════
# 4. GUARDAR EXCEL POR SEGMENTO
# ══════════════════════════════════════════════════════════════════════════
print('\n── 4. Exportando Excel ──')

out_xlsx = './segmentos_lamudi.xlsx'

def write_sheet_with_links(df_sheet, writer, sheet_name, score_col):
    """Escribe una hoja con la URL como hipervínculo clickeable."""
    df_out = df_sheet.copy()
    df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    url_col_idx = df_out.columns.get_loc('source_url') + 1  # 1-based

    # Encabezado en negrita
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill('solid', fgColor='2C5F8A')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Convertir URLs a hipervínculos
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col_idx)
        if cell.value and str(cell.value).startswith('http'):
            url = cell.value
            cell.value = 'Ver en Lamudi'
            cell.hyperlink = url
            cell.font = Font(color='2C5F8A', underline='single')

    # Ancho de columnas
    col_widths = {
        'name': 28, 'phone': 14, 'source_url': 18, 'address': 35,
        'property_type': 18, 'price': 14, 'colonia': 20,
        'area_m2': 10, 'bedrooms': 10, 'bathrooms': 10, 'parking': 10,
        'furnished': 10, 'has_elevator': 12, 'has_security_24h': 14,
        'has_gym': 10, 'has_pool': 10, 'amenidades_count': 14,
        'notes': 35, score_col: 12,
    }
    for i, col in enumerate(df_out.columns, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 14)


with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:

    # Hoja resumen
    resumen = pd.DataFrame({
        'Segmento':        ['Estudiantes', 'Tercera Edad', 'Extranjeros/Chinos'],
        'Top propiedades': [len(top_est), len(top_3ra), len(top_ext)],
        'Score máximo':    [top_est['score_estudiante'].max(),
                            top_3ra['score_tercera_edad'].max(),
                            top_ext['score_extranjero'].max()],
        'Precio mediano':  [top_est['price'].median(),
                            top_3ra['price'].median(),
                            top_ext['price'].median()],
        'Con teléfono':    [top_est['phone'].notna().sum(),
                            top_3ra['phone'].notna().sum(),
                            top_ext['phone'].notna().sum()],
    })
    resumen.to_excel(writer, sheet_name='Resumen', index=False)

    # Una hoja por segmento con hipervínculos
    write_sheet_with_links(top_est, writer, 'Estudiantes',        'score_estudiante')
    write_sheet_with_links(top_3ra, writer, 'Tercera_Edad',       'score_tercera_edad')
    write_sheet_with_links(top_ext, writer, 'Extranjeros_Chinos', 'score_extranjero')

    # Hoja con todos los datos enriquecidos
    df.drop(columns=['description'], errors='ignore').to_excel(
        writer, sheet_name='Todos_los_leads', index=False)

print(f'  ✓ Excel guardado: {out_xlsx}')

# CSV ranking consolidado
ranks = pd.concat([
    top_est[['name','phone','source_url','property_type','price','colonia',
             'area_m2','bedrooms','score_estudiante']].assign(segmento='Estudiantes'),
    top_3ra[['name','phone','source_url','property_type','price','colonia',
             'area_m2','bedrooms','score_tercera_edad']].rename(
                 columns={'score_tercera_edad':'score'}).assign(segmento='Tercera_Edad'),
    top_ext[['name','phone','source_url','property_type','price','colonia',
             'area_m2','bedrooms','score_extranjero']].rename(
                 columns={'score_extranjero':'score'}).assign(segmento='Extranjeros'),
], ignore_index=True)

ranks.to_csv('./segmentos_lamudi_ranks.csv', index=False)
print('  ✓ CSV ranking guardado: segmentos_lamudi_ranks.csv')

# ══════════════════════════════════════════════════════════════════════════
# RESUMEN FINAL
# ══════════════════════════════════════════════════════════════════════════
print('\n' + '='*60)
print('RESUMEN FINAL')
print('='*60)
print(f'Total leads analizados : {len(df):,}')
print()
for seg, top, scol in [
    ('Estudiantes',        top_est, 'score_estudiante'),
    ('Tercera edad',       top_3ra, 'score_tercera_edad'),
    ('Extranjeros/Chinos', top_ext, 'score_extranjero'),
]:
    con_tel = top['phone'].notna().sum()
    print(f'{seg}:')
    print(f'  Score máximo     : {top[scol].max()}')
    print(f'  Precio mediano   : ${top["price"].median():,.0f} MXN/mes')
    print(f'  Leads con tel    : {con_tel}/30')
    print(f'  Top colonia      : {top["colonia"].value_counts().index[0]}')
    print()
print(f'Figuras en: {OUT}')
print(f'Excel    : {out_xlsx}')
print('='*60)
